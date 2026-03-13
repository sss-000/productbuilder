from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


COLUMN_ALIASES = {
    "date": ["date", "날짜", "학습일", "오답일", "풀이일", "일자"],
    "subject": ["subject", "과목", "영역", "분류", "카테고리"],
    "unit": ["unit", "단원", "챕터", "chapter", "파트", "파트명"],
    "question": ["question", "문제", "문항", "문제내용", "질문"],
    "source": ["source", "출처", "교재", "시험명", "자료"],
    "my_answer": ["my_answer", "내답", "내 답", "작성답", "오답", "학생답"],
    "correct_answer": ["correct_answer", "정답", "모범답안", "답", "정답지"],
    "reason": ["reason", "오답이유", "틀린이유", "실수이유", "원인", "메모"],
    "note": ["note", "해설", "오답노트", "정리", "복습내용", "비고"],
}

OUTPUT_COLUMNS = [
    "학습일",
    "과목",
    "단원",
    "문제",
    "출처",
    "내 답",
    "정답",
    "오답 유형",
    "오답 이유",
    "복습 메모",
]


def normalize_name(value: str) -> str:
    return re.sub(r"[^0-9a-z가-힣]+", "", str(value).strip().lower())


def find_input_file(explicit_path: str | None) -> Path:
    if explicit_path:
        file_path = Path(explicit_path).expanduser().resolve()
        if not file_path.exists():
            raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {file_path}")
        return file_path

    candidates = sorted(
        path
        for path in Path.cwd().glob("*.xlsx")
        if not path.name.startswith("~$") and "정리본" not in path.stem
    )
    if not candidates:
        raise FileNotFoundError("현재 폴더에 처리할 xlsx 파일이 없습니다.")
    return candidates[0].resolve()


def load_sheet(file_path: Path, sheet_name: str | None) -> pd.DataFrame:
    return pd.read_excel(file_path, sheet_name=sheet_name or 0)


def map_columns(df: pd.DataFrame) -> dict[str, str]:
    normalized_columns = {normalize_name(column): column for column in df.columns}
    mapped: dict[str, str] = {}

    for target, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            match = normalized_columns.get(normalize_name(alias))
            if match:
                mapped[target] = match
                break

    return mapped


def pick_value(row: pd.Series, mapped_columns: dict[str, str], key: str) -> str:
    column = mapped_columns.get(key)
    if not column:
        return ""

    value = row.get(column, "")
    if pd.isna(value):
        return ""
    return str(value).strip()


def classify_reason(reason: str, note: str, my_answer: str, correct_answer: str) -> str:
    text = f"{reason} {note}".strip().lower()
    rules = [
        ("개념 부족", ["개념", "이론", "공식", "정의", "암기"]),
        ("계산 실수", ["계산", "실수", "부호", "산수"]),
        ("조건 누락", ["조건", "누락", "못봄", "안읽", "지문"]),
        ("시간 부족", ["시간", "촉박", "급함"]),
        ("단순 오기", ["오기", "오타", "마킹"]),
    ]

    for label, keywords in rules:
        if any(keyword in text for keyword in keywords):
            return label

    if my_answer and correct_answer and my_answer == correct_answer:
        return "정답 일치"
    if text:
        return "기타"
    return "미분류"


def build_clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    mapped_columns = map_columns(df)

    rows: list[dict[str, str]] = []
    for _, row in df.iterrows():
        record = {
            "학습일": pick_value(row, mapped_columns, "date"),
            "과목": pick_value(row, mapped_columns, "subject") or "미분류",
            "단원": pick_value(row, mapped_columns, "unit"),
            "문제": pick_value(row, mapped_columns, "question"),
            "출처": pick_value(row, mapped_columns, "source"),
            "내 답": pick_value(row, mapped_columns, "my_answer"),
            "정답": pick_value(row, mapped_columns, "correct_answer"),
            "오답 이유": pick_value(row, mapped_columns, "reason"),
            "복습 메모": pick_value(row, mapped_columns, "note"),
        }
        record["오답 유형"] = classify_reason(
            record["오답 이유"],
            record["복습 메모"],
            record["내 답"],
            record["정답"],
        )
        rows.append(record)

    clean_df = pd.DataFrame(rows, columns=[col for col in OUTPUT_COLUMNS if col != "오답 유형"])
    clean_df.insert(7, "오답 유형", [row["오답 유형"] for row in rows])

    clean_df = clean_df.replace(r"^\s*$", pd.NA, regex=True)
    clean_df["학습일"] = pd.to_datetime(clean_df["학습일"], errors="coerce").dt.strftime("%Y-%m-%d")
    clean_df["학습일"] = clean_df["학습일"].fillna("")

    sort_columns = [column for column in ["과목", "학습일", "단원"] if column in clean_df.columns]
    clean_df = clean_df.sort_values(sort_columns, na_position="last").fillna("")
    clean_df.index = range(1, len(clean_df) + 1)
    return clean_df


def build_summary_dataframe(clean_df: pd.DataFrame) -> pd.DataFrame:
    if clean_df.empty:
        return pd.DataFrame(columns=["과목", "오답 수", "주요 오답 유형"])

    summary = (
        clean_df.groupby("과목", dropna=False)
        .agg(
            오답수=("문제", "count"),
            주요오답유형=("오답 유형", lambda series: series.value_counts().index[0] if not series.empty else ""),
        )
        .reset_index()
    )
    summary.columns = ["과목", "오답 수", "주요 오답 유형"]
    return summary.sort_values(["오답 수", "과목"], ascending=[False, True])


def autosize_and_style(output_path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 40)

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(output_path)


def build_output_path(input_path: Path, explicit_output: str | None) -> Path:
    if explicit_output:
        return Path(explicit_output).expanduser().resolve()
    return input_path.with_name(f"{input_path.stem}_정리본.xlsx")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="엑셀 오답노트를 정리해서 새 문서로 저장합니다.")
    parser.add_argument("input", nargs="?", help="입력 xlsx 파일 경로")
    parser.add_argument("-s", "--sheet", help="읽을 시트 이름")
    parser.add_argument("-o", "--output", help="출력 xlsx 파일 경로")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = find_input_file(args.input)
    output_path = build_output_path(input_path, args.output)

    source_df = load_sheet(input_path, args.sheet)
    clean_df = build_clean_dataframe(source_df)
    summary_df = build_summary_dataframe(clean_df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        clean_df.to_excel(writer, sheet_name="오답정리", index_label="번호")
        summary_df.to_excel(writer, sheet_name="요약", index=False)

    autosize_and_style(output_path)

    print(f"입력 파일: {input_path}")
    print(f"출력 파일: {output_path}")
    print(f"정리된 오답 수: {len(clean_df)}")


if __name__ == "__main__":
    main()
